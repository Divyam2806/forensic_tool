from pathlib import Path

# ─────────────────────────────────────────────
# Helper: safely get a value from a dict-like
# object without crashing if the key is missing
# ─────────────────────────────────────────────
def _safe_get(obj, key, default="unavailable"):
    try:
        val = obj.get(key) if hasattr(obj, 'get') else getattr(obj, key, None)
        return str(val).strip() if val else default
    except:
        return default


# ─────────────────────────────────────────────
# PDF handler
# Uses pypdf to read the document information
# dictionary stored in the PDF's cross-reference table
# ─────────────────────────────────────────────
def _extract_pdf(path: str) -> dict:
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        meta   = reader.metadata

        return {
            "format":       "PDF",
            "page_count":   len(reader.pages),
            "encrypted":    reader.is_encrypted,
            "author":       _safe_get(meta, "/Author"),
            "creator_tool": _safe_get(meta, "/Creator"),   # Software that made the original file
            "producer":     _safe_get(meta, "/Producer"),  # Software that converted it to PDF
            "title":        _safe_get(meta, "/Title"),
            "subject":      _safe_get(meta, "/Subject"),
            "created":      _safe_get(meta, "/CreationDate"),
            "modified":     _safe_get(meta, "/ModDate"),

            # Forensic relevance: creator_tool and producer together
            # can tell you exactly what software and version made this PDF.
            # A PDF claiming to be a scanned document but showing
            # "Microsoft Word 2019" as creator is immediately suspicious.
        }
    except Exception as e:
        return {"format": "PDF", "error": str(e)}


# ─────────────────────────────────────────────
# DOCX handler
# DOCX files are ZIP archives containing XML files.
# python-docx reads the core.xml and app.xml files
# inside that ZIP — these store document properties.
# ─────────────────────────────────────────────
def _extract_docx(path: str) -> dict:
    try:
        from docx import Document
        doc  = Document(path)
        core = doc.core_properties

        return {
            "format":          "DOCX",
            "author":          _safe_get(core, "author"),
            "last_modified_by":_safe_get(core, "last_modified_by"),
            "created":         str(core.created)  if core.created  else "unavailable",
            "modified":        str(core.modified) if core.modified else "unavailable",
            "revision":        _safe_get(core, "revision"),      # How many times it was saved
            "editing_time":    _safe_get(core, "revision"),
            "title":           _safe_get(core, "title"),
            "subject":         _safe_get(core, "subject"),
            "keywords":        _safe_get(core, "keywords"),
            "description":     _safe_get(core, "description"),
            "category":        _safe_get(core, "category"),
            "company":         _safe_get(core, "identifier"),

            # Forensic relevance: author and last_modified_by are
            # pulled from Windows user account settings at the time
            # of saving — they often contain the real username even
            # if the file was renamed or moved afterward.
            # revision count shows how many times the file was saved —
            # a "draft" with 47 revisions tells a different story.
        }
    except Exception as e:
        return {"format": "DOCX", "error": str(e)}


# ─────────────────────────────────────────────
# XLSX handler
# Same ZIP/XML structure as DOCX — openpyxl reads
# the workbook properties from the XML inside
# ─────────────────────────────────────────────
def _extract_xlsx(path: str) -> dict:
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
        prop = wb.properties

        return {
            "format":           "XLSX",
            "author":           _safe_get(prop, "creator"),
            "last_modified_by": _safe_get(prop, "lastModifiedBy"),
            "created":          str(prop.created)  if prop.created  else "unavailable",
            "modified":         str(prop.modified) if prop.modified else "unavailable",
            "title":            _safe_get(prop, "title"),
            "subject":          _safe_get(prop, "subject"),
            "keywords":         _safe_get(prop, "keywords"),
            "description":      _safe_get(prop, "description"),
            "category":         _safe_get(prop, "category"),
            "sheet_names":      wb.sheetnames,
            "sheet_count":      len(wb.sheetnames),
        }
    except Exception as e:
        return {"format": "XLSX", "error": str(e)}


# ─────────────────────────────────────────────
# PPTX handler
# Same ZIP/XML structure — python-docx's core_properties
# works for PPTX too since they share the same OOXML standard
# ─────────────────────────────────────────────
def _extract_pptx(path: str) -> dict:
    try:
        from pptx import Presentation
        prs  = Presentation(path)
        core = prs.core_properties

        return {
            "format":           "PPTX",
            "author":           _safe_get(core, "author"),
            "last_modified_by": _safe_get(core, "last_modified_by"),
            "created":          str(core.created)  if core.created  else "unavailable",
            "modified":         str(core.modified) if core.modified else "unavailable",
            "revision":         _safe_get(core, "revision"),
            "title":            _safe_get(core, "title"),
            "slide_count":      len(prs.slides),
        }
    except Exception as e:
        return {"format": "PPTX", "error": str(e)}


# ─────────────────────────────────────────────
# Image handler — JPEG, PNG, TIFF, BMP, WebP
# Pillow reads EXIF data embedded in image files.
# EXIF (Exchangeable Image File Format) is a standard
# for storing camera and location metadata inside images.
# ─────────────────────────────────────────────
def _extract_image(path: str) -> dict:
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        img  = Image.open(path)
        info = {
            "format":     img.format,
            "mode":       img.mode,       # Color mode — RGB, RGBA, L (grayscale) etc
            "width_px":   img.width,
            "height_px":  img.height,
        }

        # Not all images have EXIF — PNG often doesn't, JPEG almost always does
        raw_exif = img._getexif() if hasattr(img, '_getexif') else None

        if raw_exif:
            exif = {TAGS.get(k, k): v for k, v in raw_exif.items()}

            info["camera_make"]  = exif.get("Make",             "unavailable")
            info["camera_model"] = exif.get("Model",            "unavailable")
            info["software"]     = exif.get("Software",         "unavailable")
            info["datetime"]     = exif.get("DateTime",         "unavailable")
            info["shutter"]      = exif.get("ExposureTime",     "unavailable")
            info["aperture"]     = exif.get("FNumber",          "unavailable")
            info["iso"]          = exif.get("ISOSpeedRatings",  "unavailable")
            info["focal_length"] = exif.get("FocalLength",      "unavailable")
            info["flash"]        = exif.get("Flash",            "unavailable")

            # ── GPS extraction ─────────────────────────────────────────
            # GPS data is nested inside EXIF as a sub-dictionary.
            # The coordinates are stored in degrees/minutes/seconds
            # format and need to be converted to decimal degrees
            # (the format Google Maps uses) to be readable.
            gps_info = exif.get("GPSInfo")
            if gps_info:
                gps = {GPSTAGS.get(k, k): v for k, v in gps_info.items()}

                def dms_to_decimal(dms, ref):
                    # DMS = Degrees, Minutes, Seconds
                    # Each is a rational number (numerator/denominator tuple)
                    try:
                        d = float(dms[0])
                        m = float(dms[1])
                        s = float(dms[2])
                        decimal = d + m / 60 + s / 3600
                        # Southern latitude and Western longitude are negative
                        if ref in ["S", "W"]:
                            decimal = -decimal
                        return round(decimal, 6)
                    except:
                        return "unavailable"

                lat = dms_to_decimal(
                    gps.get("GPSLatitude",  [0, 0, 0]),
                    gps.get("GPSLatitudeRef",  "N")
                )
                lon = dms_to_decimal(
                    gps.get("GPSLongitude", [0, 0, 0]),
                    gps.get("GPSLongitudeRef", "E")
                )

                info["gps_latitude"]  = lat
                info["gps_longitude"] = lon
                info["gps_altitude"]  = str(gps.get("GPSAltitude", "unavailable"))
                info["gps_maps_link"] = (
                    f"https://maps.google.com/?q={lat},{lon}"
                    if isinstance(lat, float) else "unavailable"
                )
            else:
                info["gps"] = "no GPS data in EXIF"
        else:
            info["exif"] = "no EXIF data found"

        return info

    except Exception as e:
        return {"format": "IMAGE", "error": str(e)}


# ─────────────────────────────────────────────
# HEIC handler
# HEIC is Apple's image format used by iPhones.
# pillow-heif registers itself as a Pillow plugin,
# after that we can read it like any other image.
# ─────────────────────────────────────────────
def _extract_heic(path: str) -> dict:
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
        # Once registered, Pillow can open HEIC files
        # so we just delegate to the image handler
        return _extract_image(path)
    except Exception as e:
        return {"format": "HEIC", "error": str(e)}


# ─────────────────────────────────────────────
# Media handler — MP4, MOV, AVI, MKV, MP3
# pymediainfo wraps the MediaInfo application
# which reads container-level metadata from
# virtually any audio or video format
# ─────────────────────────────────────────────
def _extract_media(path: str) -> dict:
    try:
        from pymediainfo import MediaInfo
        media = MediaInfo.parse(path)

        result = {"format": Path(path).suffix.upper().strip(".")}

        for track in media.tracks:
            if track.track_type == "General":
                # General track contains file-level metadata
                result["duration_ms"]       = track.duration
                result["file_size_bytes"]   = track.file_size
                result["created"]           = str(track.encoded_date or "unavailable")
                result["modified"]          = str(track.tagged_date  or "unavailable")
                result["encoder"]           = str(track.writing_application or "unavailable")
                result["title"]             = str(track.title       or "unavailable")
                result["artist"]            = str(track.performer   or "unavailable")
                result["album"]             = str(track.album       or "unavailable")
                result["track_number"]      = str(track.track_name  or "unavailable")
                result["genre"]             = str(track.genre       or "unavailable")
                result["overall_bitrate"]   = str(track.overall_bit_rate or "unavailable")

            elif track.track_type == "Video":
                result["video_codec"]       = str(track.codec_id    or "unavailable")
                result["resolution"]        = f"{track.width}x{track.height}"
                result["frame_rate"]        = str(track.frame_rate  or "unavailable")
                result["bit_depth"]         = str(track.bit_depth   or "unavailable")

            elif track.track_type == "Audio":
                result["audio_codec"]       = str(track.codec_id    or "unavailable")
                result["audio_channels"]    = str(track.channel_s   or "unavailable")
                result["audio_sample_rate"] = str(track.sampling_rate or "unavailable")
                result["audio_bitrate"]     = str(track.bit_rate    or "unavailable")

        return result

    except Exception as e:
        return {"format": "MEDIA", "error": str(e)}

def _extract_audio(path: str) -> dict:
    try:
        from mutagen import File as MutagenFile

        audio = MutagenFile(path, easy=True)
        if audio is None:
            return {"format": "AUDIO", "error": "Unsupported or corrupt audio file"}

        info = {
            "format":       Path(path).suffix.upper().strip("."),
            "duration_sec": round(audio.info.length, 2) if hasattr(audio.info, 'length') else "unavailable",
            "bitrate":      getattr(audio.info, 'bitrate', "unavailable"),
            "sample_rate":  getattr(audio.info, 'sample_rate', "unavailable"),
            "channels":     getattr(audio.info, 'channels', "unavailable"),
        }

        # easy=True gives simplified tag names across formats
        tags = audio.tags or {}
        info["title"]  = tags.get("title",  ["unavailable"])[0]
        info["artist"] = tags.get("artist", ["unavailable"])[0]
        info["album"]  = tags.get("album",  ["unavailable"])[0]
        info["genre"]  = tags.get("genre",  ["unavailable"])[0]
        info["date"]   = tags.get("date",   ["unavailable"])[0]
        info["track_number"] = tags.get("tracknumber", ["unavailable"])[0]

        return info

    except Exception as e:
        return {"format": "AUDIO", "error": str(e)}

# ─────────────────────────────────────────────
# Text handler — TXT, CSV, LOG
# No embedded metadata standard exists for plain text.
# We derive what we can: encoding, line count,
# a content preview for quick triage.
# ─────────────────────────────────────────────
def _extract_text(path: str) -> dict:
    try:
        # Try UTF-8 first, fall back to latin-1 which never fails
        # since it maps every byte to a character
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines    = f.readlines()
            encoding = "utf-8"
        except UnicodeDecodeError:
            with open(path, "r", encoding="latin-1") as f:
                lines    = f.readlines()
            encoding = "latin-1"

        return {
            "format":        Path(path).suffix.upper().strip("."),
            "encoding":      encoding,
            "line_count":    len(lines),
            "word_count":    sum(len(l.split()) for l in lines),
            "char_count":    sum(len(l) for l in lines),
            "preview":       [l.rstrip() for l in lines[:5]],  # First 5 lines
        }
    except Exception as e:
        return {"format": "TEXT", "error": str(e)}


# ─────────────────────────────────────────────
# Unknown file handler
# For any format we don't recognise, read the
# first 16 bytes — these are called "magic bytes"
# and most file formats have a unique signature
# that identifies the true file type regardless
# of what the extension says.
# ─────────────────────────────────────────────
def _extract_unknown(path: str) -> dict:
    try:
        with open(path, "rb") as f:
            magic = f.read(16)

        return {
            "format":      "UNKNOWN",
            "magic_bytes": magic.hex().upper(),
            "magic_ascii": "".join(chr(b) if 32 <= b < 127 else "." for b in magic),
            # Common signatures to recognise manually:
            # 25 50 44 46 → %PDF → actual PDF with wrong extension
            # 50 4B 03 04 → PK.. → ZIP / DOCX / XLSX / PPTX
            # FF D8 FF    → JPEG
            # 89 50 4E 47 → PNG
            # 52 61 72 21 → RAR archive
        }
    except Exception as e:
        return {"format": "UNKNOWN", "error": str(e)}


# ─────────────────────────────────────────────
# Format dispatcher — maps extensions to handlers
# ─────────────────────────────────────────────
EXTENSION_MAP = {
    ".pdf":  _extract_pdf,
    ".docx": _extract_docx,
    ".xlsx": _extract_xlsx,
    ".pptx": _extract_pptx,
    ".jpg":  _extract_image,
    ".jpeg": _extract_image,
    ".png":  _extract_image,
    ".tiff": _extract_image,
    ".tif":  _extract_image,
    ".bmp":  _extract_image,
    ".webp": _extract_image,
    ".heic": _extract_heic,
    ".mp4":  _extract_media,
    ".mov":  _extract_media,
    ".avi":  _extract_media,
    ".mkv":  _extract_media,
    ".mp3":  _extract_audio,
    ".wav":  _extract_audio,
    ".flac": _extract_audio,
    ".ogg":  _extract_audio,
    ".m4a":  _extract_audio,
    ".aiff": _extract_audio,
    ".wma":  _extract_audio,
    ".txt":  _extract_text,
    ".csv":  _extract_text,
    ".log":  _extract_text,
}


# ─────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────
def extract_format_metadata(path: str) -> dict:
    """
    Given a file path, detect its format and extract
    all embedded metadata from inside the file.
    Returns a dict with format-specific metadata fields.
    """
    path = Path(path).resolve()

    if not path.exists():
        return {"error": f"Path does not exist: {path}"}

    if not path.is_file():
        return {"error": f"Not a file: {path}"}

    ext     = path.suffix.lower()
    handler = EXTENSION_MAP.get(ext, _extract_unknown)

    result          = handler(str(path))
    result["path"]  = str(path)
    result["name"]  = path.name

    return result


# ─────────────────────────────────────────────
# Directory scanner — same pattern as fs_metadata
# ─────────────────────────────────────────────
def extract_from_directory(directory: str, recursive: bool = True, max_files: int = 1000) -> dict:
    """
    Walk a directory and call extract() on every file found.
    """
    directory = Path(directory).resolve()

    if not directory.exists():
        return {"error": f"Directory does not exist: {directory}"}

    results = []
    errors  = []
    count   = 0

    walk = directory.rglob("*") if recursive else directory.glob("*")

    for item in walk:
        if count >= max_files:
            print(f"[!] Reached max_files limit ({max_files}). Stopping.")
            break
        if item.is_file():
            meta = extract_format_metadata(str(item))
            if "error" in meta:
                errors.append(meta)
            else:
                results.append(meta)
            count += 1

    return {
        "scanned_directory": str(directory),
        "total_files_found": len(results),
        "total_errors":      len(errors),
        "files":             results,
        "errors":            errors,
    }